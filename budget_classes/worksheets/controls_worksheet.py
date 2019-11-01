"""BalanceWorksheet class.
    """
from openpyxl.styles import Alignment, Color, fills, PatternFill,  Font, NamedStyle

from .budget_worksheet import BudgetWorksheet
from workbook_init_scripts.border_presets import title_borders


class ControlWorksheet(BudgetWorksheet):
    """A modified openpyxl worksheet class
        """
    def __init__(self, parent, title=None):
        BudgetWorksheet.__init__(self, parent, title)
        self.set_title_header()
        self.style_title_header()

    # ------------------------- IncomeWorksheet Class Methods -----------------------------
    # --------------------------- Methods for sheet headers ---------------------
    def set_title_header(self):
        """
            Merges cells at top of sheet to from title area.
            Add title to merged cells.
        """
        self.merge_cells('A1:F3')
        self.cell(row=1, column=1, value='Control')

    def style_title_header(self):
        """
        Creates a NamedStyle object and applies the style
        to the sheet title.

        Adds border sheet header.
        """

        title_style = NamedStyle(name='control_title')
        title_style.font = Font(size=20, bold=True, underline='single',)
        title_style.alignment = Alignment(horizontal='center', vertical='center')
        title_style.fill = PatternFill(fgColor=Color("AA00FF"), fill_type=fills.FILL_SOLID)
        c = self.cell(row=1, column=1)
        c.style = title_style

        title_borders(self, 'A1:F3')

    # --------------------------- Methods for sheet headers ---------------------
    def set_field_labels(self):
        """
        Sets the name and column widths of the field headers
        in the sheet.

        To be implemented by one of four subclasses.
        """

        print("To be implemented by one of four subclasses.")

    def style_field_labels(self):
        """
        Styles the fields headers in the sheet.
            -font
            -border
            -alignment

        To be implemented by one of four subclasses.
        """

        print("To be implemented by one of four subclasses.")