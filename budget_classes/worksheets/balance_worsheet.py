"""BalanceWorksheet class.
    """
from openpyxl.styles import Alignment, Color, fills, PatternFill,  Font, NamedStyle

from .budget_worksheet import BudgetWorksheet
from workbook_init_scripts.border_presets import title_border, f_label_border


class BalanceWorksheet(BudgetWorksheet):
    """
    A modified openpyxl worksheet class.

    When instantiated, a styled template is created.
    Displays the calculated totals of the income,
    expenses, and the balance between the two.

    These values are to be derived from the worksheets
    'IncomeWorksheet' and 'ExpenseWorksheet' located within
    the same workbook.
    """
    def __init__(self, parent, title=None):
        BudgetWorksheet.__init__(self, parent, title)
        self.set_title_header()
        self.set_field_labels()
        self.set_field_placeholders()

    # ------------------------- IncomeWorksheet Class Methods -----------------------------
    # --------------------------- Methods for sheet headers ---------------------
    def set_title_header(self):
        """
        Set alignment, font, fill, and border properties
        for the title cell.

        Merge the first cell with range of cells that will
        comprise the sheet title header.
        """

        title_style = NamedStyle(name='balance_title')
        title_style.font = Font(size=20, bold=True, underline='single',)
        title_style.alignment = Alignment(horizontal='center', vertical='center')
        title_style.fill = PatternFill(fgColor=Color("6666CC"), fill_type=fills.FILL_SOLID)
        title_style.border = title_border

        self['A1'].value = 'Balance'
        self['A1'].style = title_style
        self.merge_cells('A1:D3')

    # --------------------------- Methods for sheet headers ---------------------
    def set_field_labels(self):
        """
        Sets the alignment, border, font, and value properties
        of the cells that display the field labels.

        Sets the column width of the columns containing the fields.

        Merges the cells that comprise each field.

        -Expense
        -Income
        -Balance
        """

        # Field label ranges
        fh_ranges = {
            'Expense': 'A4:B4',
            'Income': 'C4:D4',
            'Balance': 'B8:C8'}

        # Label alignment style
        label_style = NamedStyle(name='balance_fields')
        label_style.alignment = Alignment(horizontal='center', vertical='center')
        label_style.border = f_label_border
        label_style.font = Font(bold=True)

        # Set column widths
        for column in ('A', 'B', 'C', 'D'):
            self.column_dimensions[column].width = 15

        # Set field label values, merge applicable cells, add borders
        for index, (key, value) in enumerate(fh_ranges.items()):
            # Get first cell of field label
            first = value.split(':')[0]

            # Set field label, style and merge
            self[first].value = key
            self[first].style = label_style
            self.merge_cells(value)

    def set_field_placeholders(self):
        """Unique to the balance_worksheet subclass.

        Sets the alignment, border, value and number format
        properties of the cells that display the field
        placeholders.

        Set the field values to 0 and display currency format."""
        ph_ranges = {'Expenses': 'A5:B6',
                     'Income': 'C5:D6',
                     'Balance': 'B9:C10'}

        # Create NamedStyle to apply to the cell
        ph_style = NamedStyle(name="ph_style")
        ph_style.alignment = Alignment(horizontal='center', vertical='center')
        ph_style.border = f_label_border
        ph_style.number_format = '$#,##0.00'

        # Set default placeholder values (0), apply style, merge cells
        for index, value in enumerate(ph_ranges.values()):
            cell_addr = value.split(':')[0]
            self[cell_addr].value = 0
            self[cell_addr].style = ph_style
            self.merge_cells(value)
