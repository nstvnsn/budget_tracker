"""
BudgetWorkbook class
Inherits from openpyxl.workbook.workbook.Workbook
"""

from openpyxl.workbook.workbook import (Workbook,
                                        ReadOnlyWorkbookException,
                                        WriteOnlyWorksheet)
from openpyxl.worksheet.copier import WorksheetCopy

from budget_classes.worksheets.balance_worksheet import BalanceWorksheet
from budget_classes.worksheets.control_worksheet import ControlWorksheet
from budget_classes.worksheets.expense_worksheet import ExpenseWorksheet
from budget_classes.worksheets.income_worksheet import IncomeWorksheet


class BudgetWorkbook(Workbook):

    def create_sheet(self, title=None, index=None):
        """Overrides parent method to create custom
        sheets based on title.

        :param title: title of the sheet
        :type title: str
        :param index: optional position at which the sheet will be inserted
        :type index: int

        """

        if title == 'Income':
            new_ws = IncomeWorksheet(parent=self, title=title)
        elif title == 'Balance':
            new_ws = BalanceWorksheet(parent=self, title=title)
        elif title == 'Expense':
            new_ws = ExpenseWorksheet(parent=self, title=title)
        elif title == 'Control':
            new_ws = ControlWorksheet(parent=self, title=title)

        self._add_sheet(sheet=new_ws)
        return new_ws

    def initialize_budget_workbook(self):
        """Creates a new workbook with 4 sheets,
        one for each derived Worksheet class.

        :return:
        """
        active = self.active
        self.create_sheet('Expense')
        self.create_sheet('Income')
        self.create_sheet('Balance')
        self.create_sheet('Control')
        self.remove_sheet(active)

        self.save('./wb_budget.xlsx')
