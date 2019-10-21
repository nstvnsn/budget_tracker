"""Worksheet Class inheriting from openpyxl.worksheet.worksheet"""
from openpyxl.cell.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, \
    NamedStyle
from openpyxl.styles.borders import BORDER_MEDIUM, BORDER_THIN

from workbook_init_scripts import (set_borders,
                                   set_alignments,
                                   set_cell_format,
                                   sides as pss)


class BudgetWorksheet(Worksheet):
    """BudgetWorksheet inherits from openpyxl.worksheet.worksheet.Worksheet"""

    # ----------------------- Overridden Class Methods -----------------------------
    def __init__(self, parent, title=None):
        super().__init__(parent, title)
        self._setup()


    def _setup(self):
        super()._setup()
        self.set_sheet_header()

        if self.title == 'Expenses':
            self.set_expense_field_headers()

        elif self.title == 'Income':
            self.set_income_field_headers()

        elif self.title == 'Income':
            self.set_income_field_headers()
        else:
            self.set_balance_field_headers()



    # ---------------------------- New Class Methods -----------------------------
    # --------------------------- Methods for sheet headers ---------------------
    def set_sheet_header(self):
        """
            Merges cells at top of each sheet to from title area.

            :param ws: excel worksheet object
            :param title: title of worksheet object
            """

        # Merges the correct cells on each sheet
        if self.title not in ('Balance', 'Controls'):
            self.merge_cells('A1:C3')
        elif self.title == 'Balance':
            self.merge_cells('A1:D3')
        else:
            self.merge_cells('A1:F3')

        self.cell(row=1, column=1, value=self.title)

    def style_titles(self):
        """
        Creates a NamedStyle object and applies the style
        to each sheet header.

        Each sheet header is given it's own fill pattern

        Adds border to each sheet header

        :param wb: Workbook containing the sheet
        :return:
        """

        title_style = NamedStyle(name='title_font')
        title_style.font = Font(size=20, bold=True, underline='single')
        title_style.alignment = Alignment(horizontal='center', vertical='center')

        hp = None

        if self.title == 'Expenses':
            print('SELF TITLE:', self.title)
            hp = ('CC6666', 'solid')
        elif self.title == 'Income':
            hp = ('66CC66', 'solid')
        elif self.title == 'Balance':
            hp = ('6666CC', 'solid')
        elif self.title == 'Controls':
            hp = ('AA00FF', 'solid')
        else:
            raise ValueError('worksheet has an invalid title:\n'
                             'Expenses, Income, Balance, Controls\n'
                             'Showing - Title: ' + self.title)

        # assigns first cell in merged cell to c to apply a style
        c = Cell(self, row=1, column=1)
        print(type(c))


        # bd object for use with setting border style
        bd = Border(
            left=pss.side_medium,
            right=pss.side_medium,
            top=pss.side_thick,
            bottom=pss.side_thick
        )

        # Applies border to each header, using bd object
        if self.title not in ('Controls', 'Balance'):
            max_col = self[4][-1].column_letter  # uses fields instead of title as...
            title_range = f'A1:{max_col}3'        # ...merged cells don't have this attribute
        else:
            title_range = f'A1:F3'
        set_borders.border_range(self, cell_range=title_range, border=bd)

    # --------------------------- Methods for field headers ---------------------
    def set_expense_field_headers(self):
        """
        Date
        Purchase
        Expenditure

        :param:
        :return:
        """

        # Date field - date
        c = self.cell(row=4, column=1, value='Date')
        self.column_dimensions[c.column_letter].width = 15

        # Purchase field - string
        c = self.cell(row=4, column=2, value='Purchase')
        self.column_dimensions[c.column_letter].width = 25

        # Expenditure field - float
        c = self.cell(row=4, column=3, value='Expenditure')
        self.column_dimensions[c.column_letter].width = 15

    def set_income_field_headers(self):

        # Date field - date
        c = self.cell(row=4, column=1, value='Date')
        self.column_dimensions[c.column_letter].width = 15

        # Source field - string
        c = self.cell(row=4, column=2, value='Source')
        self.column_dimensions[c.column_letter].width = 25

        # Income field - float
        c = self.cell(row=4, column=3, value='Income')
        self.column_dimensions[c.column_letter].width = 15

    def set_balance_field_headers(self):

        # Field header ranges
        f_ranges = {'Expenses': 'A4:B4',
                    'Income': 'C4:D4',
                    'Balance': 'B8:C8'}

        for column in ('A', 'B', 'C', 'D'):
            self.column_dimensions[column].width = 15

        for index, (key, value) in enumerate(f_ranges.items()):
            # set field labels
            # uses first cell address of range to set value
            self[value.split(':')[0]].value = key

            # Merge field header cells
            # Merge takes place after field title is set
            self.merge_cells(value)