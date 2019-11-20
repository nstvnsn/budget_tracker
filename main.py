"""The main script for the budget_tracker_excel project.

    The program first checks if the wb_budget.xlsx workbook exists,
    if not one is created and setup. Otherwise the existing workbook
    will be opened.

    The new or existing workbook is stored in variable "wb" for later
    use.
"""

from openpyxl import load_workbook

from budget_classes.budget_workbook import (BudgetWorkbook,
                                            ExpenseWorksheet,
                                            IncomeWorksheet,
                                            BalanceWorksheet)
from budget_classes.records.expense_record import ExpenseRecord

F_NAME = './wb_budget.xlsx'


def access_workbook():

    try:
        wb = load_workbook(F_NAME)
        print(f'Workbook "{F_NAME[2:]}" loaded')
        wb.active = wb['Expense']
        ws = wb.active
        print(ws)
        record = ExpenseRecord('01/01/2109', 'Food', 9000)
        record.add_new_record(ws)
        return wb

    except FileNotFoundError:
        print('No workbook found in program directory', end='\n\n')
        print(f'Creating new workbook "{F_NAME[2:]}"')
        wb = BudgetWorkbook()
        wb.initialize_budget_workbook()

        return wb


def main():
    wb = access_workbook()
    wb.save(F_NAME)












if __name__ == '__main__':
    main()
