"""
New beginnings.
This file will develop into a means of creating or overwriting
.xlsx workbook that this program will use for output.

The program first checks if the wb_budget.xlsx workbook exists,
if not one is created and setup. Otherwise the existing workbook
will be opened.

Returns wb object, program can then populate with records from database
"""

from openpyxl import load_workbook

from workbook.workbook import BudgetWorkbook

F_NAME = './excel/wb_budget.xlsx'


def access_workbook():

    try:
        wb = load_workbook(F_NAME)
        print(f'Workbook "{F_NAME[8:]}" loaded')
        return wb

    except FileNotFoundError:
        print('No workbook found in program directory', end='\n\n')
        print(f'Creating new workbook "{F_NAME[8:]}"')
        wb = BudgetWorkbook()
        wb.initialize_budget_workbook()

    return wb


if __name__ == "__main__":
    access_workbook()
