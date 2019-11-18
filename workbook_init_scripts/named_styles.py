from openpyxl.styles import NamedStyle, Alignment, Font, PatternFill, Color, fills

from workbook_init_scripts.border_presets import f_label_border, title_border



# --------------- Title header styles ---------------
# Balance header
if NamedStyle(name='balance_title'):  # Checks for existence of named style before creating new one
    title_style_b = 'balance_title'
else:
    title_style_b = NamedStyle(name='balance_title')
    title_style_b.font = Font(size=20, bold=True, underline='single', )
    title_style_b.alignment = Alignment(horizontal='center', vertical='center')
    title_style_b.fill = PatternFill(fgColor=Color("6666CC"), fill_type=fills.FILL_SOLID)
    title_style_b.border = title_border

# Expense header
if NamedStyle(name='expense_title'):
    title_style_e = 'expense_title'
else:
    title_style_e = NamedStyle(name='expense_title')
    title_style_e.font = Font(size=20, bold=True, underline='single', )
    title_style_e.alignment = Alignment(horizontal='center', vertical='center')
    title_style_e.fill = PatternFill(fgColor=Color("CC6666"), fill_type=fills.FILL_SOLID)
    title_style_e.border = title_border

# Income header
if NamedStyle(name='income_title'):
    title_style_i = 'income_title'
else:
    title_style_i = NamedStyle(name='income_title')
    title_style_i.font = Font(size=20, bold=True, underline='single', )
    title_style_i.alignment = Alignment(horizontal='center', vertical='center')
    title_style_i.fill = PatternFill(fgColor=Color("66CC66"), fill_type=fills.FILL_SOLID)
    title_style_i.border = title_border

# Control header
if NamedStyle(name='control_title'):
    title_style_c = 'control_title'
else:
    title_style_c = NamedStyle(name='control_title')
    title_style_c.font = Font(size=20, bold=True, underline='single', )
    title_style_c.alignment = Alignment(horizontal='center', vertical='center')
    title_style_c.fill = PatternFill(fgColor=Color("AA00FF"), fill_type=fills.FILL_SOLID)
    title_style_c.border = title_border


# --------------- Field Styles ---------------
# Field label style
if NamedStyle(name='field_labels'):
    label_style = 'field_labels'
else:
    label_style = NamedStyle(name='field_labels')
    label_style.alignment = Alignment(horizontal='center', vertical='center')
    label_style.border = f_label_border
    label_style.font = Font(bold=True)

# Field placeholder style w/ currency format
if NamedStyle(name="ph_style"):
    ph_style = "ph_style"
else:
    ph_style = NamedStyle(name="ph_style")
    ph_style.alignment = Alignment(horizontal='center', vertical='center')
    ph_style.border = f_label_border
    ph_style.number_format = '$#,##0.00'

