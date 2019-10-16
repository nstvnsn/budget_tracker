from openpyxl.styles.numbers import BUILTIN_FORMATS, NumberFormat


def set_to_currency(ws, cell=None, cell_range=None):

    if cell:
        cell_ref = ws[cell]
        cell_ref.number_format = BUILTIN_FORMATS[5]     # '$#,##0.00'

    if cell_range:
        col_letter = cell_range[0]  # from string like 'A2:A22'
        for index, row in ws[cell_range]:
            ws[f'{col_letter}{index}'].number_format = '$#,##0.00'




