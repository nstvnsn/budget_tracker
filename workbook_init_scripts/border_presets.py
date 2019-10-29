from openpyxl.styles import borders, Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.styles.borders import BORDER_THICK, BORDER_MEDIUM, BORDER_THIN


def border_cell(cell, border=Border()):
    cell.border += border


def border_range(ws, cell_range, border=Border()):
    """
        Apply border styles to a range of cells as if they were a single cell.

        :param ws: Excel worksheet instance
        :param cell_range: An excel range to style (e.g. A1:F20)
        :param border: An openpyxl Border
        """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    rows = ws[cell_range]
    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        left_c = row[0]
        right_c = row[-1]
        left_c.border = left_c.border + left
        right_c.border = right_c.border + right


def title_borders(ws, cell_range):
    t_border = Border(
        top=Side(style='thick', color='000000'),
        bottom=Side(style='thick', color='000000'),
        left=Side(style='thick', color='000000'),
        right=Side(style='thick', color='000000')
    )
    border_range(ws, cell_range, t_border)


def new_record_borders(ws, cell_range, border=Border()):

    side_medium = Side(border_style=BORDER_MEDIUM, color='000000')
    side_thin = Side(border_style=BORDER_THIN, color='000000')

    left = Border(left=side_medium)
    right = Border(right=side_medium)
    right_inner = Border(right=side_thin)
    bottom = Border(bottom=side_thin)

    rows = ws[cell_range]
    for row in rows:
        for cell in row:
            if cell == row[0]:
                cell.border += left + right_inner
            elif cell == row[-1]:
                cell.border += right
            else:
                cell.border += right_inner

            if cell in list(ws)[0]:
                cell.border += bottom
