from openpyxl.styles import Border, Side, PatternFill, Font, Alignment, \
    NamedStyle
from openpyxl.styles.borders import BORDER_MEDIUM, BORDER_THIN


from workbook_init_scripts import (set_borders,
                                   set_alignments,
                                   set_cell_format,
                                   sides as pss)


# --------------------- Sheet Headers ---------------------
def style_titles(wb):
    """
    Creates a NamedStyle object and applies the style
    to each sheet header.

    Each sheet header is given it's own fill pattern

    Adds border to each sheet header

    :param wb: Workbook containing the sheet
    :return:
    """

    title_fonts = NamedStyle(name='title_font')
    title_fonts.font = Font(size=20, bold=True, underline='single')
    title_fonts.alignment = Alignment(horizontal='center', vertical='center')

    worksheets = wb.worksheets
    for index, sheet in enumerate(worksheets):
        hp = (('CC6666', 'solid'),  # Expenses
              ('66CC66', 'solid'),  # Income
              ('6666CC', 'solid'),  # Balance
              ('AA00FF', 'solid'))  # Controls

        # assigns first cell in merged cell to c to apply a style
        c = sheet.cell(row=1, column=1)
        c.style = title_fonts
        c.fill = PatternFill(fgColor=hp[index][0], fill_type=hp[index][1])

        # bd object for use with setting border style
        bd = Border(
            left=pss.side_medium,
            right=pss.side_medium,
            top=pss.side_thick,
            bottom=pss.side_thick
        )

        # Applies border to each header, using bd object
        if sheet is not wb['Controls']:
            max_col = sheet[4][-1].column_letter  # uses fields instead of title as...
            title_range = f'A1:{max_col}3'        # ...merged cells don't have this attribute
        else:
            title_range = f'A1:F3'
        set_borders.border_range(sheet, cell_range=title_range, border=bd)


# --------------------- Field Headers ---------------------
def add_field_header_borders(wb):
    """
        Apply medium borders to the field headers on each sheet within the
        workbook application's workbook. Applies thin borders as dividers
        between cells.

        :param wb: the application's wb
        :return: None
        """

    side = Side(border_style=BORDER_MEDIUM, color='00000000')
    side_thin = Side(border_style=BORDER_THIN, color='00000000')
    bd = Border(
        left=side,
        right=side,
        bottom=side,
        top=side
    )
    bd_divider = Border(right=side_thin)

    for sheet in wb:
        if sheet in [wb['Balance'], wb['Controls']]:
            # Balance is special, different layout than first 2 sheets
            # Controls has no fields
            if sheet is wb['Balance']:
                # Expenses and Income, bd wraps both
                field_range = 'A4:D4'
                set_borders.border_range(sheet, cell_range=field_range, border=bd)

                # Balance
                field_range = 'B8:C8'
                set_borders.border_range(sheet, cell_range=field_range, border=bd)

                # Set dividing border between field headers
                set_borders.border_cell(sheet[4][1], border=bd_divider)
            continue

        # sets border around field header area
        max_col = sheet[4][-1].column_letter  # uses fields instead of title as...
        field_range = f'A4:{max_col}4'        # ...merged cells don't have this attribute
        set_borders.border_range(sheet, cell_range=field_range, border=bd)

        # sets dividers between cells
        for cell in sheet[4]:  # Cell in row 4 of given sheet
            if cell is not sheet[4][-1]:
                set_borders.border_cell(cell, border=bd_divider)


def align_field_headers(wb):
    """
    Center align field header titles
    :param wb:
    :return:
    """

    align = Alignment(vertical='center', horizontal='center')
    for sheet in wb:
        if sheet in [wb['Balance'], wb['Controls']]:
            if sheet is wb['Controls']:
                continue

            f_ranges = {'Expenses': 'A4:B4',
                        'Income': 'C4:D4',
                        'Balance': 'B8:C8'}
            for index, value in enumerate(f_ranges.values()):
                # Center align field labels in Balance sheet
                set_alignments.align_range(sheet, cell_range=value, alignment=align)
            continue

        max_col = sheet[4][-1].column_letter  # uses fields instead of title as...
        field_range = f'A4:{max_col}4'  # ...merged cells don't have this attribute
        set_alignments.align_range(sheet, cell_range=field_range, alignment=align)


def set_balance_fv_placeholders(ws):
    # Placeholder ranges
    ph_ranges = {'expenses': 'A5:B6',
                 'income': 'C5:D6',
                 'balance': 'B9:C10'}

    align = Alignment(vertical='center', horizontal='center')

    bd_expenses = Border(
        left=pss.side_medium,
        bottom=pss.side_medium,
        right=pss.side_thin
    )
    bd_income = Border(
        right=pss.side_medium,
        bottom=pss.side_medium,
    )
    bd_balance = Border(
        left=pss.side_medium,
        right=pss.side_medium,
        bottom=pss.side_medium
    )

    bd = [bd_expenses,
          bd_income,
          bd_balance]

    for index, value in enumerate(ph_ranges.values()):
        # Set ph values to 0
        ws[value.split(':')[0]].value = 0

        # Merge ph cells
        ws.merge_cells(value)

        # Center align cells
        set_alignments.align_range(ws, cell_range=value, alignment=align)

        # Set format to currency
        set_cell_format.set_to_currency(ws, value.split(':')[0])

        # Set Borders
        set_borders.border_range(ws, value, border=bd[index])
