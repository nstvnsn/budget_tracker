from openpyxl.styles import Alignment
from . import set_alignments

"""
    For each of the 4 sheets:
        -Creates sheet header
        -Creates appropriate fields for each sheet
    """


# --------------------------- Sheet Headers ---------------------------
def set_sheet_headers(wb):
    """
        Merges cells at top of each sheet to from title area.

        :param wb: excel workbook instance
        """

    worksheets = wb.worksheets

    for index, sheet in enumerate(worksheets):

        # Merges the correct cells on each sheet
        if sheet not in (wb['Balance'], wb['Controls']):
            sheet.merge_cells('A1:C3')
        elif sheet == wb['Balance']:
            sheet.merge_cells('A1:D3')
        else:
            sheet.merge_cells('A1:F3')

        hp = ('Expenses', 'Income', 'Balance', 'Controls')
        sheet.cell(row=1, column=1, value=hp[index])


# --------------------- Field Headers ---------------------
def set_expense_field_headers(wb):
    """
    Date
    Purchase
    Expenditure

    :param wb:
    :return:
    """
    ws = wb.worksheets[0]

    # Date field - date
    c = ws.cell(row=4, column=1, value='Date')
    ws.column_dimensions[c.column_letter].width = 15

    # Purchase field - string
    c = ws.cell(row=4, column=2, value='Purchase')
    ws.column_dimensions[c.column_letter].width = 25

    # Expenditure field - float
    c = ws.cell(row=4, column=3, value='Expenditure')
    ws.column_dimensions[c.column_letter].width = 15


def set_income_field_headers(wb):
    ws = wb.worksheets[1]

    # Date field - date
    c = ws.cell(row=4, column=1, value='Date')
    ws.column_dimensions[c.column_letter].width = 15

    # Source field - string
    c = ws.cell(row=4, column=2, value='Source')
    ws.column_dimensions[c.column_letter].width = 25

    # Income field - float
    c = ws.cell(row=4, column=3, value='Income')
    ws.column_dimensions[c.column_letter].width = 15


def set_balance_field_headers(wb):
    ws = wb.worksheets[2]

    # Field header ranges
    f_ranges = {'Expenses': 'A4:B4',
                'Income': 'C4:D4',
                'Balance': 'B8:C8'}

    align = Alignment(vertical='center', horizontal='center')

    for column in ('A', 'B', 'C', 'D'):
        ws.column_dimensions[column].width = 15

    for index, (key, value) in enumerate(f_ranges.items()):
        # set field labels
        # uses first cell address of range to set value
        ws[value.split(':')[0]].value = key

        # Merge field header cells
        # Merge takes place after field title is set
        ws.merge_cells(value)


def set_field_headers(wb):
    set_expense_field_headers(wb)
    set_income_field_headers(wb)
    set_balance_field_headers(wb)
